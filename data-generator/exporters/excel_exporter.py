"""
Excel Exporter Module

Exports data to Excel format with:
- Multiple sheets
- Formatted headers
- Summary statistics
- Some intentional formatting issues (merged cells)
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


def export_to_excel(dfs_dict, filename, output_dir='output'):
    """
    Export multiple DataFrames to Excel with multiple sheets
    
    Args:
        dfs_dict: Dictionary of {sheet_name: DataFrame}
        filename: Output filename (without extension)
        output_dir: Output directory path
        
    Returns:
        Path to exported file
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    filepath = output_path / f"{filename}.xlsx"
    
    # Write DataFrames to Excel
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for sheet_name, df in dfs_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"✅ Exported to {filepath}")
    print(f"   Sheets: {list(dfs_dict.keys())}")
    
    return filepath


def export_formatted_excel(dfs_dict, filename, output_dir='output'):
    """
    Export to Excel with formatting (headers, colors)
    
    This creates a more "business user friendly" Excel file
    """
    print(f"\n📄 Exporting formatted Excel: {filename}.xlsx")
    
    filepath = export_to_excel(dfs_dict, filename, output_dir)
    
    # Apply formatting
    wb = load_workbook(filepath)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format header row
        header_fill = PatternFill(start_color="E8896A", end_color="E8896A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filepath)
    
    return filepath


def export_messy_excel(dfs_dict, filename, output_dir='output'):
    """
    Export to Excel with intentional formatting issues
    
    Simulates receiving data from business users who:
    - Merge cells
    - Add extra header rows
    - Use inconsistent formatting
    """
    print(f"\n📄 Exporting messy Excel: {filename}.xlsx")
    
    filepath = export_to_excel(dfs_dict, filename, output_dir)
    
    # Apply messy formatting
    wb = load_workbook(filepath)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Insert a title row at the top (shifts data down)
        ws.insert_rows(1)
        ws['A1'] = f"{sheet_name} Data Export"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Merge some cells in title row
        ws.merge_cells('A1:D1')
        
        # Add a blank row
        ws.insert_rows(2)
        
        # Now the actual headers are in row 3
        # Add some color to random cells (messy)
        random_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Color a few random cells
        if ws.max_row > 5:
            ws['B5'].fill = random_fill
            ws['C7'].fill = random_fill
    
    wb.save(filepath)
    
    return filepath


def create_summary_sheet(df, sheet_name='Summary'):
    """
    Create a summary statistics sheet
    
    Args:
        df: DataFrame to summarize
        sheet_name: Name for the summary sheet
        
    Returns:
        DataFrame with summary statistics
    """
    summary_data = []
    
    # Basic stats
    summary_data.append({
        'Metric': 'Total Records',
        'Value': len(df)
    })
    
    # Numeric column stats
    numeric_cols = df.select_dtypes(include=['number']).columns
    
    for col in numeric_cols:
        summary_data.append({
            'Metric': f'{col} - Total',
            'Value': df[col].sum()
        })
        summary_data.append({
            'Metric': f'{col} - Average',
            'Value': df[col].mean()
        })
        summary_data.append({
            'Metric': f'{col} - Min',
            'Value': df[col].min()
        })
        summary_data.append({
            'Metric': f'{col} - Max',
            'Value': df[col].max()
        })
    
    # Date range if timestamp exists
    if 'timestamp' in df.columns:
        summary_data.append({
            'Metric': 'Date Range Start',
            'Value': df['timestamp'].min()
        })
        summary_data.append({
            'Metric': 'Date Range End',
            'Value': df['timestamp'].max()
        })
    
    return pd.DataFrame(summary_data)


if __name__ == "__main__":
    # Test Excel exporter
    import pandas as pd
    
    # Test data
    products_df = pd.DataFrame({
        'sku': ['LM-FOOD-001', 'CC-BEV-001'],
        'name': ['Lucky Me Pancit Canton', 'Coca-Cola 1.5L'],
        'price': [15.00, 65.00],
        'cost_price': [12.00, 52.00]
    })
    
    sales_df = pd.DataFrame({
        'transaction_id': ['TXN001', 'TXN002'],
        'product_name': ['Lucky Me Pancit Canton', 'Coca-Cola 1.5L'],
        'quantity': [2, 1],
        'total_amount': [30.00, 65.00],
        'timestamp': [datetime.now(), datetime.now()]
    })
    
    summary_df = create_summary_sheet(sales_df)
    
    # Export formatted
    export_formatted_excel({
        'Products': products_df,
        'Sales': sales_df,
        'Summary': summary_df
    }, 'test_formatted')
    
    # Export messy
    export_messy_excel({
        'Sales': sales_df
    }, 'test_messy')
